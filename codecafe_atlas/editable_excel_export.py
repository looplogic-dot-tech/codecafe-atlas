from __future__ import annotations

from pathlib import Path
from typing import Iterable, Mapping, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EXPORT_HEADERS = [
    "Nombre del edificio",
    "Calle o avenida",
    "Número exterior",
    "Colonia",
    "Código postal",
    "Ciudad",
    "Estado",
    "País",
    "Piso",
    "Dependencia, juzgado, tribunal u oficina",
    "CTA / encargado de la dependencia",
    "Teléfono de la dependencia",
    "Correo de la dependencia",
    "Oficina / grupo de trabajo",
    "Tipo de equipo",
    "Marca del equipo",
    "Modelo del equipo",
    "Número de serie",
    "Número de inventario",
    "Usuario del equipo",
    "Dirección IP",
    "Nombre de red / hostname",
    "Estado del equipo",
    "Observaciones del equipo",
]

EXPORT_KEYS = [
    "building_name", "street", "exterior_number", "colony", "postal_code",
    "city", "state", "country", "floor", "dependency_name", "cta_name",
    "dependency_phone", "dependency_email", "office_name", "equipment_type",
    "brand", "model", "serial_number", "inventory_number", "assigned_user",
    "ip_address", "hostname", "status", "equipment_notes",
]


def export_editable_excel(path: Path, rows: Iterable[Mapping[str, Any]]) -> int:
    """Write a flat, migration-assistant-friendly XLSX with one row per equipment."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario Atlas"
    ws.append(EXPORT_HEADERS)

    count = 0
    for row in rows:
        values = dict(row)
        ws.append(["" if values.get(key) is None else str(values.get(key)) for key in EXPORT_KEYS])
        count += 1

    # A simple first-row header is intentional: Atlas Data Bridge can detect it
    # without requiring any Atlas-only metadata or hidden database identifiers.
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 34

    widths = {
        1: 30, 2: 28, 3: 15, 4: 24, 5: 15, 6: 20, 7: 20, 8: 16,
        9: 12, 10: 44, 11: 30, 12: 22, 13: 30, 14: 28, 15: 18, 16: 18,
        17: 22, 18: 24, 19: 22, 20: 28, 21: 18, 22: 26, 23: 18, 24: 38,
    }
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.number_format = "@"

    # Keep serials, inventory numbers, IPs and hostnames as literal text even
    # when Excel would otherwise infer a numeric/date format.
    for col in (18, 19, 21, 22):
        letter = get_column_letter(col)
        for cell in ws[letter][1:]:
            cell.number_format = "@"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return count
