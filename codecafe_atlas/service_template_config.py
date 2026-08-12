from __future__ import annotations

import json
import re
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QAbstractItemView,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QComboBox,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
)

from .service_document_generator import (
    SERVICE_REQUIRED_PLACEHOLDERS,
    SERVICE_SUPPORTED_PLACEHOLDERS,
    validate_service_template,
)


FIELD_LABELS = {
    "{{REPORTE_DGTI}}": "Reporte DGTI",
    "{{REPORTE_PRESTADOR}}": "Reporte del prestador",
    "{{FECHA_REPORTE}}": "Fecha del reporte",
    "{{HORA_REPORTE}}": "Hora del reporte",
    "{{RESPONSABLE_EQUIPO}}": "Responsable del equipo",
    "{{DEPENDENCIA}}": "Dependencia",
    "{{DOMICILIO}}": "Domicilio",
    "{{CIUDAD_ESTADO}}": "Ciudad / Estado",
    "{{TELEFONO_DEPENDENCIA}}": "Teléfono de dependencia",
    "{{VALIDADOR}}": "Servidor público que valida",
    "{{CARGO_VALIDADOR}}": "Cargo del validador",
    "{{TELEFONO_VALIDADOR}}": "Teléfono del validador",
    "{{MOV_SUSTITUCION}}": "Movimiento: Sustitución",
    "{{MOV_ACTUALIZACION}}": "Movimiento: Actualización",
    "{{MOV_REUBICACION}}": "Movimiento: Reubicación",
    "{{MOV_INCREMENTO}}": "Movimiento: Incremento",
    "{{MOV_DISMINUCION}}": "Movimiento: Disminución",
    "{{FALLA_REPORTADA}}": "Falla reportada",
    "{{TIPO_EQUIPO}}": "Tipo de equipo",
    "{{MARCA}}": "Marca",
    "{{MODELO}}": "Modelo",
    "{{NUMERO_SERIE}}": "Número de serie",
    "{{NUMERO_INVENTARIO}}": "Número de inventario",
    "{{FECHA_DIAGNOSTICO}}": "Fecha de diagnóstico",
    "{{HORA_DIAGNOSTICO}}": "Hora de diagnóstico",
    "{{FECHA_SOLUCION}}": "Fecha de solución",
    "{{HORA_SOLUCION}}": "Hora de solución",
    "{{DIAGNOSTICO}}": "Diagnóstico",
    "{{SOLUCION}}": "Solución / servicio",
    "{{OBSERVACIONES}}": "Observaciones",
    "{{TECNICO}}": "Técnico",
    "{{RESPONSABLE_FIRMA}}": "Responsable / firma",
}

CELL_RE = re.compile(r"^[A-Z]{1,3}[1-9][0-9]*$")


def load_template_settings(path: Path) -> dict[str, Any]:
    try:
        if not path.exists():
            return {}
        payload = json.loads(path.read_text(encoding="utf-8"))
        return payload if isinstance(payload, dict) else {}
    except (OSError, ValueError, TypeError):
        return {}


def save_template_settings(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_cell_list(value: str) -> list[str]:
    cells: list[str] = []
    for raw in re.split(r"[,;\s]+", str(value or "").strip().upper()):
        if not raw:
            continue
        if not CELL_RE.fullmatch(raw):
            raise ValueError(f"Referencia de celda no válida: {raw}")
        if raw not in cells:
            cells.append(raw)
    return cells


def workbook_sheet_names(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


class ServiceTemplateConfigDialog(QDialog):
    """Configures the service-certificate template without changing Atlas data schema."""

    def __init__(
        self,
        parent,
        *,
        included_template: Path,
        config_path: Path,
        managed_template_path: Path,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("Configuración inicial del generador de cédulas — CodeCafe Atlas")
        self.resize(900, 720)
        self.included_template = Path(included_template)
        self.config_path = Path(config_path)
        self.managed_template_path = Path(managed_template_path)
        self._settings = load_template_settings(self.config_path)
        self._selected_source: Path | None = None

        root = QVBoxLayout(self)
        intro = QLabel(
            "Selecciona la plantilla que utilizará Atlas. Los placeholders {{CAMPO}} se "
            "detectan automáticamente. Si tu plantilla no usa placeholders, puedes indicar "
            "directamente una o varias celdas para cada campo (por ejemplo M8 o A15,K20)."
        )
        intro.setWordWrap(True)
        root.addWidget(intro)

        template_row = QHBoxLayout()
        self.template_edit = QLineEdit()
        self.template_edit.setReadOnly(True)
        use_included = QPushButton("Usar plantilla incluida")
        choose_custom = QPushButton("Elegir plantilla propia…")
        template_row.addWidget(self.template_edit, 1)
        template_row.addWidget(use_included)
        template_row.addWidget(choose_custom)
        root.addLayout(template_row)

        sheet_row = QHBoxLayout()
        sheet_row.addWidget(QLabel("Hoja que Atlas debe completar:"))
        self.sheet_combo = QComboBox()
        sheet_row.addWidget(self.sheet_combo, 1)
        root.addLayout(sheet_row)

        self.status = QLabel("")
        self.status.setWordWrap(True)
        root.addWidget(self.status)

        helper = QLabel(
            "Mapeo avanzado (opcional). Deja una celda vacía cuando la plantilla ya contiene "
            "el placeholder correspondiente. Una misma entrada puede indicar varias celdas "
            "separadas por comas."
        )
        helper.setWordWrap(True)
        root.addWidget(helper)

        tokens = sorted(SERVICE_SUPPORTED_PLACEHOLDERS, key=lambda token: FIELD_LABELS.get(token, token))
        self.table = QTableWidget(len(tokens), 4)
        self.table.setHorizontalHeaderLabels(["Campo Atlas", "Placeholder", "Celda(s)", "Requerido"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setStretchLastSection(True)
        saved_map = self._settings.get("cell_map") or {}
        for row, token in enumerate(tokens):
            label_item = QTableWidgetItem(FIELD_LABELS.get(token, token))
            label_item.setFlags(label_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            token_item = QTableWidgetItem(token)
            token_item.setFlags(token_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            raw_cells = saved_map.get(token, [])
            if isinstance(raw_cells, str):
                raw_cells = [raw_cells]
            cells_item = QTableWidgetItem(", ".join(str(cell) for cell in raw_cells if str(cell).strip()))
            required_item = QTableWidgetItem("Sí" if token in SERVICE_REQUIRED_PLACEHOLDERS else "")
            required_item.setFlags(required_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row, 0, label_item)
            self.table.setItem(row, 1, token_item)
            self.table.setItem(row, 2, cells_item)
            self.table.setItem(row, 3, required_item)
        self.table.resizeColumnsToContents()
        root.addWidget(self.table, 1)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.button(QDialogButtonBox.StandardButton.Save).setText("Guardar configuración")
        root.addWidget(buttons)

        use_included.clicked.connect(self.select_included)
        choose_custom.clicked.connect(self.select_custom)
        buttons.accepted.connect(self.accept_configuration)
        buttons.rejected.connect(self.reject)

        current = str(self._settings.get("template_path") or "").strip()
        if current and Path(current).exists():
            self._selected_source = Path(current)
            self.template_edit.setText(current)
            self.inspect_selected_template()
        else:
            self.select_included()

    def select_included(self) -> None:
        self._selected_source = self.included_template
        self.template_edit.setText(str(self.included_template))
        self.inspect_selected_template()

    def select_custom(self) -> None:
        selected, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar plantilla Excel",
            str(Path.home()),
            "Libro de Excel (*.xlsx)",
        )
        if selected:
            self._selected_source = Path(selected)
            self.template_edit.setText(selected)
            self.inspect_selected_template()

    def inspect_selected_template(self) -> None:
        path = self._selected_source
        if path is None or not path.exists():
            self.status.setText("Selecciona una plantilla válida.")
            return
        try:
            sheets = workbook_sheet_names(path)
            previous_sheet = str(self._settings.get("sheet_name") or "").strip()
            self.sheet_combo.blockSignals(True)
            self.sheet_combo.clear()
            self.sheet_combo.addItems(sheets)
            if previous_sheet in sheets:
                self.sheet_combo.setCurrentText(previous_sheet)
            elif "Cédula de Servicio" in sheets:
                self.sheet_combo.setCurrentText("Cédula de Servicio")
            self.sheet_combo.blockSignals(False)
            found, missing, unknown = validate_service_template(path)
        except Exception as error:
            self.status.setText(f"No se pudo inspeccionar la plantilla: {error}")
            return
        parts = [f"Hojas: {', '.join(sheets)}. Placeholders reconocidos: {len(found)}."]
        if missing:
            parts.append(
                "Faltan placeholders obligatorios: " + ", ".join(sorted(missing)) +
                ". Puedes compensarlos mediante el mapeo directo de celdas."
            )
        if unknown:
            parts.append("Placeholders desconocidos: " + ", ".join(sorted(unknown)) + ".")
        self.status.setText(" ".join(parts))

    def _cell_map(self) -> dict[str, list[str]]:
        result: dict[str, list[str]] = {}
        for row in range(self.table.rowCount()):
            token = self.table.item(row, 1).text().strip()
            raw = self.table.item(row, 2).text().strip()
            if raw:
                result[token] = normalize_cell_list(raw)
        return result

    def accept_configuration(self) -> None:
        source = self._selected_source
        if source is None or not source.exists():
            QMessageBox.warning(self, "Plantilla requerida", "Selecciona una plantilla Excel válida.")
            return
        try:
            found, missing, unknown = validate_service_template(source)
            cell_map = self._cell_map()
            unresolved = sorted(token for token in missing if not cell_map.get(token))
            if unresolved:
                answer = QMessageBox.question(
                    self,
                    "Campos sin ubicación",
                    "Estos campos obligatorios no tienen placeholder ni celda asignada:\n\n"
                    + "\n".join(unresolved)
                    + "\n\nAtlas puede guardar la configuración, pero la plantilla quedará incompleta. ¿Continuar?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No,
                )
                if answer != QMessageBox.StandardButton.Yes:
                    return
            if unknown:
                answer = QMessageBox.question(
                    self,
                    "Placeholders desconocidos",
                    "La plantilla contiene placeholders que Atlas no reconoce:\n\n"
                    + "\n".join(sorted(unknown))
                    + "\n\n¿Guardar de todos modos?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No,
                )
                if answer != QMessageBox.StandardButton.Yes:
                    return

            self.managed_template_path.parent.mkdir(parents=True, exist_ok=True)
            if source.resolve() != self.managed_template_path.resolve():
                shutil.copy2(source, self.managed_template_path)
            payload = {
                "configured": True,
                "template_path": str(self.managed_template_path),
                "source_name": source.name,
                "sheet_name": self.sheet_combo.currentText().strip(),
                "cell_map": cell_map,
            }
            save_template_settings(self.config_path, payload)
        except Exception as error:
            QMessageBox.critical(self, "No se pudo guardar la configuración", str(error))
            return
        self.accept()
