from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from copy import copy


SHEET_MAP = {
    "Cédula de Servicio": "Cédula de Servicio",
    "Mantenimiento Preventivo": "Mantenimiento Preventivo",
    "Dictaminación": "Dictaminación",
}

SERVICE_TEMPLATE_SHEET = "Cédula de Servicio"
SERVICE_REQUIRED_PLACEHOLDERS = {
    "{{REPORTE_PRESTADOR}}",
    "{{FECHA_REPORTE}}",
    "{{RESPONSABLE_EQUIPO}}",
    "{{FALLA_REPORTADA}}",
    "{{MODELO}}",
    "{{NUMERO_SERIE}}",
    "{{DIAGNOSTICO}}",
    "{{SOLUCION}}",
}

SERVICE_SUPPORTED_PLACEHOLDERS = {
    "{{REPORTE_DGTI}}", "{{REPORTE_PRESTADOR}}", "{{FECHA_REPORTE}}",
    "{{HORA_REPORTE}}", "{{RESPONSABLE_EQUIPO}}", "{{DEPENDENCIA}}",
    "{{DOMICILIO}}", "{{CIUDAD_ESTADO}}", "{{TELEFONO_DEPENDENCIA}}",
    "{{VALIDADOR}}", "{{CARGO_VALIDADOR}}", "{{TELEFONO_VALIDADOR}}",
    "{{MOV_SUSTITUCION}}", "{{MOV_ACTUALIZACION}}", "{{MOV_REUBICACION}}",
    "{{MOV_INCREMENTO}}", "{{MOV_DISMINUCION}}", "{{FALLA_REPORTADA}}",
    "{{TIPO_EQUIPO}}", "{{MARCA}}", "{{MODELO}}", "{{NUMERO_SERIE}}",
    "{{NUMERO_INVENTARIO}}", "{{FECHA_DIAGNOSTICO}}",
    "{{HORA_DIAGNOSTICO}}", "{{FECHA_SOLUCION}}", "{{HORA_SOLUCION}}",
    "{{DIAGNOSTICO}}", "{{SOLUCION}}", "{{OBSERVACIONES}}",
    "{{TECNICO}}", "{{RESPONSABLE_FIRMA}}",
}


def safe_filename(value: str, fallback: str = "cedula") -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", str(value or "").strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ._")
    return cleaned or fallback


def full_address(data: dict[str, Any]) -> str:
    street_number = " ".join(
        str(part).strip()
        for part in [data.get("street", ""), data.get("exterior_number", "")]
        if str(part or "").strip()
    )
    parts = [street_number]
    colony = str(data.get("colony", "") or "").strip()
    postal_code = str(data.get("postal_code", "") or "").strip()
    building = str(data.get("building", "") or "").strip()
    floor = str(data.get("floor", "") or "").strip()
    office = str(data.get("office", "") or "").strip()
    if colony:
        parts.append(f"Col. {colony}")
    if postal_code:
        parts.append(f"C.P. {postal_code}")
    if building:
        parts.append(building)
    if floor:
        parts.append(f"Piso {floor}")
    if office:
        parts.append(office)
    return ", ".join(part for part in parts if part)


def city_state(data: dict[str, Any]) -> str:
    return " / ".join(
        str(part).strip()
        for part in [data.get("city", ""), data.get("state", "")]
        if str(part or "").strip()
    )


def write_label_value(ws, cell: str, label: str, value: Any) -> None:
    text = str(value or "").strip()
    ws[cell] = f"{label}{text}" if text else label


def write_value(ws, cell: str, value: Any) -> None:
    ws[cell] = "" if value is None else value


def fill_common_header(ws, data: dict[str, Any], document_type: str) -> None:
    if document_type == "Cédula de Servicio":
        # La clave contractual de M5 permanece fija en la plantilla revisada.
        # El reporte del prestador se presenta con la identidad CodeCafe Atlas en la plantilla
        # y el valor editable en M8.
        provider_report = str(data.get("provider_report", "") or data.get("folio", "")).strip()
        provider_report = re.sub(r"^[A-Z][A-Z0-9 ]{2,31}[-_ ]+(?=REQ[-_ ]*\d)", "", provider_report, flags=re.IGNORECASE)
        write_value(ws, "L7", data.get("dgti_report", ""))
        write_value(ws, "M8", provider_report)
        write_value(ws, "L9", data.get("report_date", ""))
        write_value(ws, "L10", data.get("report_time", ""))
    elif document_type != "Mantenimiento Preventivo":
        write_value(ws, "M5", data.get("folio", ""))
        write_value(ws, "L7", data.get("dgti_report", ""))
        write_value(ws, "L8", data.get("provider_report", ""))
        write_value(ws, "L9", data.get("report_date", ""))
        write_value(ws, "L10", data.get("report_time", ""))
    else:
        write_value(ws, "L7", data.get("report_date", ""))
        write_value(ws, "L8", data.get("report_time", ""))


def fill_responsible(ws, data: dict[str, Any], maintenance: bool = False) -> None:
    write_label_value(ws, "A15", "Nombre del responsable del equipo: ", data.get("responsible_name"))
    write_label_value(ws, "A16", "Área / Órgano Jurisdiccional: ", data.get("dependency_name"))
    write_label_value(ws, "A17", "Domicilio: ", full_address(data))

    if maintenance:
        write_label_value(ws, "A18", "Ciudad / Estado: ", city_state(data))
    else:
        write_value(ws, "C18", city_state(data).replace(" / ", " ").upper())
    write_label_value(ws, "K18", "Teléfono: ", data.get("dependency_phone"))

    if maintenance:
        label = "Nombre del servidor público designado para la validación del servicio (En caso de que aplique): "
    else:
        label = "Nombre del servidor público designado para la validación del servicio: "
    write_label_value(ws, "A19", label, data.get("validator_name"))
    if maintenance:
        write_label_value(ws, "A20", "Cargo: ", data.get("validator_role"))
    else:
        write_value(ws, "C20", data.get("validator_role", ""))
    write_label_value(ws, "K20", "Teléfono: ", data.get("validator_phone"))


def fill_equipment_row(ws, row: int, data: dict[str, Any]) -> None:
    write_value(ws, f"A{row}", data.get("equipment_type", ""))
    write_value(ws, f"E{row}", data.get("brand", ""))
    write_value(ws, f"H{row}", data.get("model", ""))
    write_value(ws, f"J{row}", data.get("serial_number", ""))
    write_value(ws, f"L{row}", data.get("inventory_number", ""))


def service_placeholder_values(data: dict[str, Any]) -> dict[str, str]:
    movement = str(data.get("movement_type", "") or "").strip().casefold()
    provider_report = str(data.get("provider_report", "") or data.get("folio", "")).strip()
    provider_report = re.sub(r"^[A-Z][A-Z0-9 ]{2,31}[-_ ]+(?=REQ[-_ ]*\d)", "", provider_report, flags=re.IGNORECASE)
    responsible_signature = str(data.get("validator_name") or data.get("responsible_name") or "").strip()
    responsible_role = str(data.get("validator_role", "") or "").strip()
    if responsible_signature and responsible_role:
        responsible_signature = f"{responsible_signature} — {responsible_role}"

    def mark(label: str) -> str:
        return "X" if movement == label.casefold() else " "

    return {
        "{{REPORTE_DGTI}}": str(data.get("dgti_report", "") or ""),
        "{{REPORTE_PRESTADOR}}": provider_report,
        "{{FECHA_REPORTE}}": str(data.get("report_date", "") or ""),
        "{{HORA_REPORTE}}": str(data.get("report_time", "") or ""),
        "{{RESPONSABLE_EQUIPO}}": str(data.get("responsible_name", "") or ""),
        "{{DEPENDENCIA}}": str(data.get("dependency_name", "") or ""),
        "{{DOMICILIO}}": full_address(data),
        "{{CIUDAD_ESTADO}}": city_state(data).replace(" / ", " ").upper(),
        "{{TELEFONO_DEPENDENCIA}}": str(data.get("dependency_phone", "") or ""),
        "{{VALIDADOR}}": str(data.get("validator_name", "") or ""),
        "{{CARGO_VALIDADOR}}": str(data.get("validator_role", "") or ""),
        "{{TELEFONO_VALIDADOR}}": str(data.get("validator_phone", "") or ""),
        "{{MOV_SUSTITUCION}}": mark("Sustitución"),
        "{{MOV_ACTUALIZACION}}": mark("Actualización"),
        "{{MOV_REUBICACION}}": mark("Reubicación"),
        "{{MOV_INCREMENTO}}": mark("Incremento"),
        "{{MOV_DISMINUCION}}": mark("Disminución"),
        "{{FALLA_REPORTADA}}": str(data.get("reported_issue", "") or ""),
        "{{TIPO_EQUIPO}}": str(data.get("equipment_type", "") or ""),
        "{{MARCA}}": str(data.get("brand", "") or ""),
        "{{MODELO}}": str(data.get("model", "") or ""),
        "{{NUMERO_SERIE}}": str(data.get("serial_number", "") or ""),
        "{{NUMERO_INVENTARIO}}": str(data.get("inventory_number", "") or ""),
        "{{FECHA_DIAGNOSTICO}}": str(data.get("diagnosis_date", data.get("report_date", "")) or ""),
        "{{HORA_DIAGNOSTICO}}": str(data.get("diagnosis_time", data.get("report_time", "")) or ""),
        "{{FECHA_SOLUCION}}": str(data.get("solution_date", data.get("report_date", "")) or ""),
        "{{HORA_SOLUCION}}": str(data.get("solution_time", data.get("report_time", "")) or ""),
        "{{DIAGNOSTICO}}": str(data.get("diagnosis", "") or ""),
        "{{SOLUCION}}": str(data.get("solution", "") or ""),
        "{{OBSERVACIONES}}": str(data.get("service_notes", "") or "Sin Observaciones"),
        "{{TECNICO}}": str(data.get("technician_name", "") or ""),
        "{{RESPONSABLE_FIRMA}}": responsible_signature,
    }


def template_placeholders(template_path: Path) -> set[str]:
    workbook = load_workbook(template_path, read_only=True, data_only=False)
    try:
        if SERVICE_TEMPLATE_SHEET not in workbook.sheetnames:
            return set()
        found: set[str] = set()
        pattern = re.compile(r"\{\{[A-Z0-9_]+\}\}")
        for row in workbook[SERVICE_TEMPLATE_SHEET].iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    found.update(pattern.findall(cell.value))
        return found
    finally:
        workbook.close()


def validate_service_template(template_path: Path) -> tuple[set[str], set[str], set[str]]:
    found = template_placeholders(Path(template_path))
    missing = SERVICE_REQUIRED_PLACEHOLDERS - found
    unknown = found - SERVICE_SUPPORTED_PLACEHOLDERS
    return found, missing, unknown


def fill_service_placeholders(ws, data: dict[str, Any]) -> None:
    replacements = service_placeholder_values(data)
    pattern = re.compile(r"\{\{[A-Z0-9_]+\}\}")
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str) or "{{" not in cell.value:
                continue
            original = cell.value
            cell.value = pattern.sub(lambda match: replacements.get(match.group(0), match.group(0)), original)


def configure_service_vertical_labels(ws) -> None:
    """Conserva verticales los dos rótulos laterales del bloque de firmas."""
    for coordinate in ("A53", "H53"):
        cell = ws[coordinate]
        alignment = copy(cell.alignment)
        alignment.textRotation = 90
        alignment.wrapText = False
        alignment.horizontal = "center"
        alignment.vertical = "center"
        cell.alignment = alignment


def configure_service_print_layout(ws) -> None:
    """Deja la cédula lista para imprimirse en una sola hoja tamaño Carta."""
    ws.print_area = "A1:M64"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.scale = None
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = False
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    # Márgenes reducidos para conservar el tamaño útil del formato sin
    # depender de ajustes manuales en Excel o LibreOffice.
    ws.page_margins.left = 0.20
    ws.page_margins.right = 0.20
    ws.page_margins.top = 0.20
    ws.page_margins.bottom = 0.20
    ws.page_margins.header = 0.10
    ws.page_margins.footer = 0.10


def fill_service_cell_map(ws, data: dict[str, Any], cell_map: dict[str, list[str]] | None = None) -> None:
    if not cell_map:
        return
    replacements = service_placeholder_values(data)
    for token, cells in cell_map.items():
        if token not in replacements:
            continue
        value = replacements[token]
        if isinstance(cells, str):
            cells = [cells]
        for coordinate in cells or []:
            coordinate = str(coordinate or "").strip().upper()
            if coordinate:
                ws[coordinate] = value


def fill_service_sheet(
    ws,
    data: dict[str, Any],
    cell_map: dict[str, list[str]] | None = None,
) -> None:
    fill_service_placeholders(ws, data)
    fill_service_cell_map(ws, data, cell_map)
    configure_service_vertical_labels(ws)
    configure_service_print_layout(ws)


def fill_maintenance_sheet(ws, data: dict[str, Any]) -> None:
    fill_common_header(ws, data, "Mantenimiento Preventivo")
    fill_responsible(ws, data, maintenance=True)
    fill_equipment_row(ws, 25, data)

    write_value(ws, "A29", data.get("solution", "") or data.get("service_notes", ""))
    operates = data.get("equipment_operates", "")
    condition = data.get("equipment_condition", "")
    write_label_value(ws, "H29", "¿El equipo opera adecuadamente? ", operates)
    write_label_value(ws, "H33", "¿El equipo presenta rayaduras o golpes? ", condition)
    write_value(ws, "A39", data.get("service_notes", "") or "Sin Observaciones")
    write_label_value(ws, "B43", "Nombre: ", data.get("technician_name", ""))
    responsible_signature = data.get("validator_name") or data.get("responsible_name")
    responsible_role = str(data.get("validator_role", "") or "").strip()
    if responsible_signature and responsible_role:
        responsible_signature = f"{responsible_signature} — {responsible_role}"
    write_label_value(ws, "I43", "Nombre y cargo del responsable del equipo: ", responsible_signature)


def fill_dictamination_sheet(ws, data: dict[str, Any]) -> None:
    fill_common_header(ws, data, "Dictaminación")
    write_label_value(ws, "A15", "Nombre del responsable del equipo: ", data.get("responsible_name"))
    write_label_value(ws, "A16", "Área / Órgano Jurisdiccional: ", data.get("dependency_name"))
    write_label_value(ws, "A17", "Domicilio: ", full_address(data))
    write_label_value(ws, "A18", "Ciudad / Estado: ", city_state(data))
    write_label_value(ws, "K18", "Teléfono: ", data.get("dependency_phone"))
    fill_equipment_row(ws, 23, data)

    write_value(ws, "A27", data.get("diagnosis", ""))
    write_value(ws, "A37", data.get("service_notes", "") or "Sin Observaciones")
    write_label_value(ws, "A41", "Nombre: ", data.get("technician_name", ""))
    write_label_value(ws, "I41", "Fecha de elaboración de la Dictaminación: ", data.get("report_date", ""))
    write_label_value(ws, "A43", "Firma: ", "")
    write_label_value(ws, "I43", "Hora de elaboración del Dictámen Técnico: ", data.get("report_time", ""))


def generate_service_document(
    template_path: Path,
    output_folder: Path,
    data: dict[str, Any],
    *,
    service_template_path: Path | None = None,
    service_cell_map: dict[str, list[str]] | None = None,
    service_sheet_name: str | None = None,
) -> Path:
    document_type = str(data.get("document_type") or "Cédula de Servicio")
    if document_type not in SHEET_MAP:
        raise ValueError(f"Tipo de documento no soportado: {document_type}")

    template_path = Path(template_path)
    output_folder = Path(output_folder)

    # La cédula de servicio usa su propia plantilla de referencia, actualizada
    # a partir del formato validado manualmente. Las hojas de mantenimiento
    # preventivo y dictaminación continúan usando el libro maestro original.
    effective_template = template_path
    if document_type == "Cédula de Servicio":
        if service_template_path is not None and Path(service_template_path).exists():
            effective_template = Path(service_template_path)
        else:
            service_template = template_path.with_name(
                "Formato de referencia - Cédula de Servicio.xlsx"
            )
            if service_template.exists():
                effective_template = service_template

    if not effective_template.exists():
        raise FileNotFoundError(f"No se encontró la plantilla:\n{effective_template}")
    output_folder.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(effective_template)
    target_sheet = SHEET_MAP[document_type]
    if document_type == "Cédula de Servicio" and service_sheet_name:
        target_sheet = str(service_sheet_name).strip()
    if target_sheet not in workbook.sheetnames:
        raise ValueError(f"La plantilla no contiene la hoja «{target_sheet}».")

    # Deja únicamente el formato seleccionado para que el archivo final sea claro.
    for sheet_name in list(workbook.sheetnames):
        if sheet_name != target_sheet:
            del workbook[sheet_name]

    worksheet = workbook[target_sheet]
    if document_type == "Cédula de Servicio":
        fill_service_sheet(worksheet, data, service_cell_map)
    elif document_type == "Mantenimiento Preventivo":
        fill_maintenance_sheet(worksheet, data)
    else:
        fill_dictamination_sheet(worksheet, data)

    folio = safe_filename(data.get("folio", ""), "sin_folio")
    serial = safe_filename(data.get("serial_number", ""), "sin_serie")
    kind = safe_filename(document_type, "cedula")
    filename = f"{kind} - {folio} - {serial}.xlsx"
    output_path = output_folder / filename

    # Evita sobreescribir silenciosamente una cédula anterior.
    if output_path.exists():
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        output_path = output_folder / f"{kind} - {folio} - {serial} - {timestamp}.xlsx"

    workbook.save(output_path)
    return output_path
