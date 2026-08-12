from __future__ import annotations

import base64
import csv
import io
import json
import os
import re
import shutil
import sys
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

from PySide6.QtCore import QObject, QUrl, Signal, Slot
from PySide6.QtGui import QDesktopServices
from PySide6.QtWebChannel import QWebChannel
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QVBoxLayout,
    QWidget,
)
from PySide6.QtWebEngineCore import QWebEngineSettings
from PySide6.QtWebEngineWidgets import QWebEngineView

from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import pytesseract
except ImportError:
    pytesseract = None

from .platform_open import open_directory_native
from .database import Database
from .paths import application_root, bundled_root, module_dir
from .ui_helpers import page_header


def _find_tesseract_executable() -> str:
    """Locate Tesseract reliably on Windows and Unix-like systems."""
    candidates: list[Path] = []

    explicit = os.environ.get("CODECAFE_ATLAS_TESSERACT_CMD") or os.environ.get("TESSERACT_CMD")
    if explicit:
        candidates.append(Path(explicit).expanduser())

    executable_name = "tesseract.exe" if sys.platform.startswith("win") else "tesseract"
    for root in (application_root(), bundled_root()):
        candidates.extend((
            root / "ocr" / "tesseract" / executable_name,
            root / "tesseract" / executable_name,
            root / executable_name,
        ))

    discovered = shutil.which("tesseract")
    if discovered:
        candidates.append(Path(discovered))

    if sys.platform.startswith("win"):
        for env_name in ("ProgramFiles", "ProgramFiles(x86)", "LOCALAPPDATA"):
            base = os.environ.get(env_name)
            if not base:
                continue
            base_path = Path(base)
            candidates.extend((
                base_path / "Tesseract-OCR" / "tesseract.exe",
                base_path / "Programs" / "Tesseract-OCR" / "tesseract.exe",
            ))

    seen: set[str] = set()
    for candidate in candidates:
        try:
            resolved = candidate.resolve()
        except OSError:
            resolved = candidate
        key = str(resolved).lower() if sys.platform.startswith("win") else str(resolved)
        if key in seen:
            continue
        seen.add(key)
        if resolved.is_file():
            return str(resolved)
    return ""


def _configure_tesseract_environment(executable: str) -> None:
    """Point pytesseract and Tesseract at the selected installation."""
    pytesseract.pytesseract.tesseract_cmd = executable
    tessdata = Path(executable).resolve().parent / "tessdata"
    if tessdata.is_dir():
        os.environ["TESSDATA_PREFIX"] = str(tessdata)


class CounterDatabaseBridge(QObject):
    ocrFinished = Signal(str, str)

    def __init__(self, database: Database):
        super().__init__()
        self.database = database
        self._ocr_executor = ThreadPoolExecutor(
            max_workers=1,
            thread_name_prefix="CodeCafe-Atlas-Counter-OCR",
        )
        self._ocr_capabilities_cache: dict | None = None

    @staticmethod
    def _response(**values) -> str:
        return json.dumps(values, ensure_ascii=False)

    def _ocr_capabilities(self) -> dict:
        if self._ocr_capabilities_cache is not None:
            return dict(self._ocr_capabilities_cache)

        executable = _find_tesseract_executable()
        available = bool(executable and pytesseract is not None)
        languages: list[str] = []
        error = ""

        if available:
            try:
                _configure_tesseract_environment(executable)
                languages = sorted(pytesseract.get_languages(config=""))
            except Exception as exception:
                available = False
                error = str(exception)

        preferred = ""
        if available:
            has_spa = "spa" in languages
            has_eng = "eng" in languages
            if has_spa and has_eng:
                preferred = "spa+eng"
            elif has_spa:
                preferred = "spa"
            elif has_eng:
                preferred = "eng"
            elif languages:
                preferred = languages[0]

        self._ocr_capabilities_cache = {
            "available": available,
            "engine": "Tesseract nativo",
            "executable": executable or "",
            "languages": languages,
            "preferred_language": preferred,
            "error": error,
        }
        return dict(self._ocr_capabilities_cache)

    @Slot(result=str)
    def ocrCapabilities(self) -> str:
        return self._response(ok=True, **self._ocr_capabilities())

    @staticmethod
    def _decode_data_url(data_url: str) -> Image.Image:
        if "," not in data_url:
            raise ValueError("La imagen OCR recibida no tiene un formato válido.")
        _, encoded = data_url.split(",", 1)
        raw = base64.b64decode(encoded, validate=False)
        image = Image.open(io.BytesIO(raw))
        image.load()
        return image.convert("RGB")

    def _recognize_native(self, data_url: str, options_payload: str) -> dict:
        capabilities = self._ocr_capabilities()
        if not capabilities["available"]:
            raise RuntimeError(capabilities.get("error") or "Tesseract nativo no está disponible.")

        try:
            options = json.loads(options_payload or "{}")
        except json.JSONDecodeError:
            options = {}

        psm = options.get("psm", 6)
        try:
            psm = max(3, min(13, int(psm)))
        except (TypeError, ValueError):
            psm = 6

        requested_language = str(
            options.get("language") or capabilities.get("preferred_language") or ""
        ).strip()
        available_languages = set(capabilities.get("languages") or [])
        requested_parts = [
            part for part in requested_language.split("+") if part in available_languages
        ]
        language = "+".join(requested_parts)

        if not language:
            if "spa" in available_languages and "eng" in available_languages:
                language = "spa+eng"
            elif "spa" in available_languages:
                language = "spa"
            elif "eng" in available_languages:
                language = "eng"
            elif available_languages:
                language = sorted(available_languages)[0]
            else:
                language = "eng"

        image = self._decode_data_url(data_url)
        config = (
            f"--oem 1 --psm {psm} "
            "-c preserve_interword_spaces=1 "
            "-c user_defined_dpi=300"
        )
        text = pytesseract.image_to_string(image, lang=language, config=config)
        return {"ok": True, "text": text or "", "engine": "native", "language": language}

    @Slot(str, str, str)
    def recognizeImage(self, request_id: str, data_url: str, options_payload: str) -> None:
        future = self._ocr_executor.submit(
            self._recognize_native, data_url, options_payload
        )

        def completed(result_future):
            try:
                response = result_future.result()
            except Exception as error:
                response = {"ok": False, "error": str(error)}
            self.ocrFinished.emit(
                request_id, json.dumps(response, ensure_ascii=False)
            )

        future.add_done_callback(completed)

    @Slot(str, result=str)
    def saveRecords(self, payload: str) -> str:
        try:
            records = json.loads(payload)
            if not isinstance(records, list):
                raise ValueError("El contenido recibido no es una lista de registros.")
            result = self.database.save_counter_records(records)
            return self._response(ok=True, **result)
        except Exception as error:
            return self._response(ok=False, error=str(error))

    @Slot(result=str)
    def listDependencies(self) -> str:
        """Return active directory dependencies for report classification only."""
        try:
            rows = self.database.dependency_choices()
            dependencies = []
            for row in rows:
                name = str(row["name"] or "").strip()
                building = str(row["building"] or "").strip()
                floor = str(row["floor"] or "").strip()
                context = " · ".join(value for value in (building, f"Piso {floor}" if floor else "") if value)
                dependencies.append({
                    "id": int(row["id"]),
                    "name": name,
                    "label": f"{name} — {context}" if context else name,
                })
            return self._response(ok=True, dependencies=dependencies)
        except Exception as error:
            return self._response(ok=False, error=str(error), dependencies=[])

    @Slot(result=str)
    def loadRecords(self) -> str:
        try:
            records = self.database.list_counter_records()
            return self._response(ok=True, records=records, count=len(records))
        except Exception as error:
            return self._response(ok=False, error=str(error))

    @Slot(str, result=str)
    def deleteRecord(self, record_uid: str) -> str:
        try:
            deleted = self.database.delete_counter_record(record_uid)
            return self._response(ok=True, deleted=deleted)
        except Exception as error:
            return self._response(ok=False, error=str(error))

    @Slot(result=str)
    def clearRecords(self) -> str:
        try:
            deleted = self.database.clear_counter_records()
            return self._response(ok=True, deleted=deleted)
        except Exception as error:
            return self._response(ok=False, error=str(error))

    @staticmethod
    def _normalise_export_value(value):
        if value is None:
            return ""
        if isinstance(value, (str, int, float, bool)):
            return value
        return str(value)

    @Slot(str, result=str)
    def exportReport(self, payload: str) -> str:
        """Save CSV or XLSX through a native Qt dialog.

        QWebEngine blob downloads are inconsistent in portable builds, so the
        HTML module sends the tabular data through QWebChannel and Python writes
        the selected file directly.
        """
        try:
            request = json.loads(payload or "{}")
            export_format = str(request.get("format") or "").lower().strip()
            suggested_name = str(request.get("filename") or "reporte_contadores").strip()
            headers = request.get("headers") or []
            rows = request.get("rows") or []
            widths = request.get("widths") or []
            fills = request.get("fills") or []

            if export_format not in {"csv", "xlsx"}:
                raise ValueError("Formato de exportación no compatible.")
            if not isinstance(headers, list) or not headers:
                raise ValueError("El reporte no contiene encabezados.")
            if not isinstance(rows, list):
                raise ValueError("Las filas del reporte no tienen un formato válido.")

            extension = ".xlsx" if export_format == "xlsx" else ".csv"
            if not suggested_name.lower().endswith(extension):
                suggested_name += extension
            file_filter = "Libro de Excel (*.xlsx)" if export_format == "xlsx" else "Archivo CSV (*.csv)"
            selected, _ = QFileDialog.getSaveFileName(
                None,
                "Guardar reporte de contadores",
                suggested_name,
                file_filter,
            )
            if not selected:
                return self._response(ok=True, cancelled=True)

            output_path = Path(selected)
            if output_path.suffix.lower() != extension:
                output_path = output_path.with_suffix(extension)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            clean_headers = [self._normalise_export_value(value) for value in headers]
            clean_rows = [
                [self._normalise_export_value(value) for value in (row if isinstance(row, list) else [])]
                for row in rows
            ]

            if export_format == "csv":
                with output_path.open("w", encoding="utf-8-sig", newline="") as stream:
                    writer = csv.writer(stream, delimiter=";", quoting=csv.QUOTE_ALL, lineterminator="\r\n")
                    writer.writerow(clean_headers)
                    writer.writerows(clean_rows)
            else:
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = str(request.get("sheetName") or "Contadores")[:31]
                # Keep the workbook in a normal, fully navigable view.
                # Frozen panes caused confusing behavior in LibreOffice Calc on
                # some Linux installations, so the report opens unfrozen.
                worksheet.freeze_panes = None
                worksheet.sheet_view.view = "normal"
                worksheet.sheet_view.showGridLines = True
                worksheet.sheet_view.showHorizontalScroll = True
                worksheet.sheet_view.showVerticalScroll = True
                worksheet.sheet_view.zoomScale = 90
                worksheet.sheet_view.zoomScaleNormal = 90
                worksheet.protection.sheet = False

                header_fill = PatternFill("solid", fgColor="074F69")
                header_font = Font(name="Aptos Narrow", size=10, bold=True, color="FFFFFF")
                thin = Side(style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                worksheet.append(clean_headers)
                for row in clean_rows:
                    padded = list(row) + [""] * max(0, len(clean_headers) - len(row))
                    worksheet.append(padded[:len(clean_headers)])

                worksheet.row_dimensions[1].height = 72
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border

                for row_index in range(2, worksheet.max_row + 1):
                    worksheet.row_dimensions[row_index].height = 20
                    for column_index in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_index, column=column_index)
                        fill_value = str(fills[column_index - 1]) if column_index - 1 < len(fills) else "FFFFFF"
                        if not re.fullmatch(r"[0-9A-Fa-f]{6}", fill_value):
                            fill_value = "FFFFFF"
                        cell.fill = PatternFill("solid", fgColor=fill_value.upper())
                        cell.font = Font(name="Aptos Narrow", size=9, color="000000")
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                for index in range(1, worksheet.max_column + 1):
                    width = widths[index - 1] if index - 1 < len(widths) else 14
                    try:
                        width = float(width)
                    except (TypeError, ValueError):
                        width = 14
                    worksheet.column_dimensions[get_column_letter(index)].width = max(8, min(width, 45))

                worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
                worksheet.print_title_rows = "1:1"
                worksheet.page_setup.orientation = "landscape"
                worksheet.page_setup.fitToWidth = 1
                worksheet.page_setup.fitToHeight = 0
                worksheet.sheet_properties.pageSetUpPr.fitToPage = True
                workbook.save(output_path)

            return self._response(ok=True, cancelled=False, path=str(output_path))
        except Exception as error:
            return self._response(ok=False, error=str(error))



class CounterRegistryPage(QWidget):
    def __init__(self, database: Database):
        super().__init__()
        self.database = database
        self.folder = module_dir("counter_registry")
        self.index_file = self.folder / "index.html"
        self._module_loaded = False

        root = QVBoxLayout(self)
        root.setContentsMargins(18, 14, 18, 18)
        root.addWidget(page_header(
            "Registro de contadores",
            "Escanea documentos, genera reportes Excel y guarda los resultados en la base común.",
        ))

        toolbar = QHBoxLayout()

        reload_button = QPushButton("Actualizar / recargar")
        reload_button.setToolTip(
            "Vuelve a cargar la versión actualmente instalada."
        )

        replace_button = QPushButton("Reemplazar módulo HTML")
        replace_button.setToolTip(
            "Selecciona una versión nueva del Registro de contadores."
        )

        browse_button = QPushButton("Buscar / abrir carpeta")
        browse_button.setToolTip(
            "Abre la carpeta donde se encuentra el módulo."
        )

        self.path_label = QLabel(str(self.index_file))
        self.path_label.setWordWrap(True)

        toolbar.addWidget(reload_button)
        toolbar.addWidget(replace_button)
        toolbar.addWidget(browse_button)
        toolbar.addWidget(self.path_label, 1)
        root.addLayout(toolbar)

        self.browser = QWebEngineView()
        settings = self.browser.settings()
        settings.setAttribute(
            QWebEngineSettings.WebAttribute.LocalContentCanAccessRemoteUrls,
            True,
        )
        settings.setAttribute(
            QWebEngineSettings.WebAttribute.LocalContentCanAccessFileUrls,
            True,
        )
        settings.setAttribute(
            QWebEngineSettings.WebAttribute.JavascriptCanAccessClipboard,
            True,
        )

        self.bridge = CounterDatabaseBridge(database)
        self.web_channel = QWebChannel(self.browser.page())
        self.web_channel.registerObject("counterDatabase", self.bridge)
        self.browser.page().setWebChannel(self.web_channel)

        root.addWidget(self.browser, 1)

        reload_button.clicked.connect(
            lambda: self.load_module(force=True)
        )
        replace_button.clicked.connect(self.replace_html)
        browse_button.clicked.connect(self.open_folder)

        self.load_module(force=True)

    def load_module(self, force: bool = False):
        if self._module_loaded and not force:
            return

        if not self.index_file.exists():
            self.index_file.write_text(
                """
                <!doctype html>
                <html lang="es">
                <meta charset="utf-8">
                <title>Módulo no encontrado</title>
                <body style="font-family:system-ui;padding:40px">
                <h1>Registro de contadores no encontrado</h1>
                <p>Use «Reemplazar módulo HTML» para seleccionar el archivo.</p>
                </body>
                </html>
                """,
                encoding="utf-8",
            )

        url = QUrl.fromLocalFile(str(self.index_file.resolve()))
        url.setQuery(f"reload={self.index_file.stat().st_mtime_ns}")
        self.browser.setUrl(url)
        self._module_loaded = True

    def replace_html(self):
        selected, _ = QFileDialog.getOpenFileName(
            self,
            "Seleccionar Registro de contadores HTML",
            str(self.folder),
            "Aplicaciones HTML (*.html *.htm)",
        )
        if not selected:
            return

        selected_path = self.index_file.__class__(selected)
        try:
            self.folder.mkdir(parents=True, exist_ok=True)

            if selected_path.resolve() != self.index_file.resolve():
                if self.index_file.exists():
                    backup = self.folder / "index_anterior.html"
                    shutil.copy2(self.index_file, backup)
                shutil.copy2(selected_path, self.index_file)

        except Exception as error:
            QMessageBox.critical(
                self,
                "No se pudo reemplazar el módulo",
                str(error),
            )
            return

        self.load_module(force=True)
        QMessageBox.information(
            self,
            "Módulo actualizado",
            "Registro de contadores fue reemplazado y recargado.\n\n"
            "La versión anterior quedó como index_anterior.html.",
        )

    def open_folder(self):
        self.folder.mkdir(parents=True, exist_ok=True)
        opened, diagnostic = open_directory_native(self.folder)
        if not opened:
            QMessageBox.warning(
                self,
                "No se pudo abrir la carpeta",
                f"Atlas no pudo abrir el administrador de archivos para:\n{self.folder.resolve()}\n\nDetalle:\n{diagnostic}",
            )
